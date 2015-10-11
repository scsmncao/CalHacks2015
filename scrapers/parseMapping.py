from openpyxl import load_workbook
import json

def parseCodeCountryMapping():
	f = open("countryHash.json", "w")
	wb = load_workbook(filename = '../data/countryList.xlsx')
	ws = wb.active
	countryHashTable = {}
	for i in range(1, 90):
		if ws['B' + str(i)].value != None:
			countryHashTable[ws['B' + str(i)].value.encode("UTF-8")] = ws['A' + str(i)].value.encode("UTF-8")
	f.write(json.dumps(countryHashTable))

parseCodeCountryMapping()